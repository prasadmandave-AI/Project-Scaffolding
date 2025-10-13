import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# ------------------------------------------------------
# Confusion Matrix Generator (Python Project)
# ------------------------------------------------------
# My objective:
# - Build a command line tool that:
#   1. Takes an Excel file as input
#   2. Reads the input data safely
#   3. Identifies all unique classifiers (disease/condition names)
#   4. Creates a new Excel file with confusion matrix structure
#   5. Adds formulas for Sensitivity, Specificity, Checks, etc.
#   6. Keeps original input data in a separate sheet
#
# Tools used:
# - pandas: for reading and processing Excel
# - openpyxl: for writing output Excel with formulas & formatting
# ------------------------------------------------------

# Final output columns for the Confusion Matrix sheet
OUTPUT_COLUMNS = [
    "condition",
    "true_Positive",
    "false_Positive",
    "true_Negative",
    "false_Negative",
    "Sensitivity",
    "Specificity",
    "Check",
    "Positive Ground Truth",
    "Negative Ground Truth",
    "Ground Truth Check",
]

# ------------------------------------------------------
# Function to read Excel input safely
# ------------------------------------------------------
def read_excel(input_path: Path) -> pd.DataFrame:
    """Reads the Excel input file and handles errors gracefully."""
    try:
        df = pd.read_excel(input_path)
        print(f"✅ Successfully read input: {input_path.name}")
        return df
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        sys.exit(1)


# ------------------------------------------------------
# Function to normalize column names
# ------------------------------------------------------
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize column names to lowercase and replace spaces with underscores.
    This makes column handling consistent and avoids key errors.
    """
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    return df


# ------------------------------------------------------
# Function to extract classifiers and build confusion matrix counts
# ------------------------------------------------------
def count_classifiers(df: pd.DataFrame) -> pd.DataFrame:
    """
    - Normalizes column names
    - Maps confusion matrix related columns (TP, FP, TN, FN)
    - Extracts unique classifiers (disease/condition labels)
    - Builds confusion matrix counts for each classifier
    """

    # Work on a copy to avoid modifying the original DataFrame
    df = normalize_columns(df.copy())

    # Dynamically identify the relevant columns in the input
    col_map = {
        "true_positive": next((c for c in df.columns if "true_positive" in c), None),
        "false_positive": next((c for c in df.columns if "false_positive" in c), None),
        "true_negative": next((c for c in df.columns if "true_negative" in c), None),
        "false_negative": next((c for c in df.columns if "false_negative" in c), None),
    }

    # Fill missing columns with empty strings to avoid errors
    present_cols = []
    for k, v in col_map.items():
        if v is None:
            print(f"⚠️ Warning: Column '{k}' not found in Excel. Skipping.")
            df[k] = ""
        else:
            df[k] = df[v].fillna("").astype(str).str.lower()
            present_cols.append(k)

    # Extract all unique classifiers
    print("🔎 Extracting classifiers from input data...")
    all_classifiers = set()

    for col_name in present_cols:
        classifiers_list = df[col_name].str.split(r'\W+').explode().str.strip()
        all_classifiers.update(classifiers_list[classifiers_list != ""].unique())

    CLASSIFIERS = sorted(list(all_classifiers))
    print(f"✅ Found {len(CLASSIFIERS)} unique classifiers.")

    # Build result confusion matrix rows
    result_rows = []
    for cond in CLASSIFIERS:
        tp = df["true_positive"].str.contains(cond, case=False, na=False, regex=False).sum()
        fp = df["false_positive"].str.contains(cond, case=False, na=False, regex=False).sum()
        tn = df["true_negative"].str.contains(cond, case=False, na=False, regex=False).sum()
        fn = df["false_negative"].str.contains(cond, case=False, na=False, regex=False).sum()

        # Append to result rows (empty strings for formula columns)
        result_rows.append([cond, tp, fp, tn, fn, "", "", "", "", "", ""])

    result_df = pd.DataFrame(result_rows, columns=OUTPUT_COLUMNS)
    return result_df


# ------------------------------------------------------
# Function to write confusion matrix + formulas to Excel
# ------------------------------------------------------
def write_excel_with_formulas(df: pd.DataFrame, input_df: pd.DataFrame, output_path: Path):
    """
    Writes:
    - Sheet 1: Confusion Matrix with formulas for Sensitivity, Specificity, etc.
    - Sheet 2: Original Input Data
    Applies formatting (center alignment).
    """

    wb = Workbook()

    # First sheet: Confusion Matrix results
    ws = wb.active
    ws.title = "Confusion Matrix"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Add formulas row-wise
    for i in range(2, len(df) + 2):
        ws[f"F{i}"] = f"=IFERROR(B{i}/(B{i}+E{i}), 0)"   # Sensitivity
        ws[f"G{i}"] = f"=IFERROR(D{i}/(D{i}+C{i}),0)"   # Specificity
        ws[f"F{i}"].number_format = '0.00%'
        ws[f"G{i}"].number_format = '0.00%'

        ws[f"H{i}"] = f"=SUM(B{i}+C{i})"               # Check = TP + FP
        ws[f"I{i}"] = f"=SUM(B{i}+E{i})"               # Positive Ground Truth
        ws[f"J{i}"] = f"=SUM(D{i}+C{i})"               # Negative Ground Truth
        ws[f"K{i}"] = f"=SUM(I{i}+J{i})"               # Ground Truth Check

    # Second sheet: Original input data
    ws_input = wb.create_sheet(title="Input Data")
    for r in dataframe_to_rows(input_df, index=False, header=True):
        ws_input.append(r)

    # Apply alignment formatting
    for ws_sheet in [ws_input, ws]:
        for row in ws_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    print(f"📂 Output file created: {output_path}")


# ------------------------------------------------------
# Main program execution
# ------------------------------------------------------
def main():
    """
    Steps:
    1. Check command line argument for input file path
    2. Read Excel input
    3. Extract classifiers & build confusion matrix
    4. Write results into a new Excel file
    """
    if len(sys.argv) < 2:
        print("Usage: python script.py <input_excel_path>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"❌ File not found: {input_path}")
        sys.exit(1)

    df = read_excel(input_path)
    output_df = count_classifiers(df)

    output_path = input_path.parent / "output_confusion_matrix.xlsx"
    write_excel_with_formulas(output_df, df, output_path)


# Run script
if __name__ == "__main__":
    main()
